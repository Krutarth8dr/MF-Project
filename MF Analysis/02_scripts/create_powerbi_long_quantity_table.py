import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]

INPUT_FILE = PROJECT_ROOT / "05_matrix" / "MASTER" / "matrix_all_amc_funds_quantity.xlsx"
OUTPUT_FILE = PROJECT_ROOT / "05_matrix" / "MASTER" / "matrix_all_amc_funds_quantity_long.xlsx"

BASE_COLUMNS = [
    "AMC",
    "Security_Name",
    "ISIN",
    "Portfolio_Date",
    "Month",
    "Industry_Rating",
]

IGNORE_COLUMNS = {
    "Total_All_AMC_Qty",
    "Total HDFC Qty",
    "Total ICICI Qty",
    "Total SBI Qty",
}

def main():
    df = pd.read_excel(INPUT_FILE)
    security_master = (
        df.groupby("ISIN")["Security_Name"]
        .apply(lambda x: sorted(set(x)))
        .reset_index()
    )

    # Convert arrays to separate columns
    names_df = pd.DataFrame(
        security_master["Security_Name"].tolist()
    )

    # Rename columns
    names_df.columns = [
        f"Name_{i+1}" for i in range(len(names_df.columns))
    ]

    # Combine with ISIN
    security_master = pd.concat(
        [security_master[["ISIN"]], names_df],
        axis=1
    )

    security_master.to_excel(
        PROJECT_ROOT / "05_matrix" / "MASTER" / "security_master.xlsx",
        index=False
    )

    df["Portfolio_Date"] = pd.to_datetime(df["Portfolio_Date"], errors="coerce")
    df = df[df["Portfolio_Date"].notna()]
    df["Month"] = format_month_text(df["Portfolio_Date"])

    fund_columns = [
        col for col in df.columns
        if col not in BASE_COLUMNS and col not in IGNORE_COLUMNS
    ]

    long_df = df.melt(
        id_vars=BASE_COLUMNS,
        value_vars=fund_columns,
        var_name="Fund_Name",
        value_name="Quantity"
    )

    long_df["Quantity"] = pd.to_numeric(long_df["Quantity"], errors="coerce").fillna(0)

    # Optional: remove zero rows to make Power BI lighter
    long_df = long_df[long_df["Quantity"] != 0]

    long_df = long_df.sort_values(
        by=["Security_Name", "Fund_Name", "Portfolio_Date"],
        ascending=[True, True, True]
    )

    long_df["Portfolio_Date"] = pd.to_datetime(long_df["Portfolio_Date"], errors="coerce")
    long_df["Month"] = format_month_text(long_df["Portfolio_Date"])
    long_df["Portfolio_Date"] = long_df["Portfolio_Date"].dt.strftime("%Y-%m-%d")

    if OUTPUT_FILE.exists():
        try:
            OUTPUT_FILE.unlink()
        except PermissionError:
            print("ERROR: Output file is open in Excel.")
            print("Please close this file and run again:")
            print(OUTPUT_FILE)
            return

    long_df.to_excel(OUTPUT_FILE, index=False)
    force_powerbi_text_columns(OUTPUT_FILE)

    print("Power BI long quantity table created.")
    print("Output:", OUTPUT_FILE)
    print("Rows:", len(long_df))
    print("Funds:", long_df["Fund_Name"].nunique())
    print("Stocks:", long_df["Security_Name"].nunique())


def format_month_text(date_series):
    month_text = pd.to_datetime(date_series, errors="coerce").dt.strftime("%b-%Y")

    # Power BI auto-detects values like Dec-2024 as dates. Using a non-breaking
    # hyphen keeps the same visual label while making the value remain text.
    return month_text.str.replace("-", "\u2011", regex=False).astype("string")


def force_powerbi_text_columns(output_file):
    workbook = load_workbook(output_file)
    worksheet = workbook.active

    text_column_names = {
        "Month",
        "Portfolio_Date",
        
    }

    text_column_indexes = []

    for cell in worksheet[1]:
        if cell.value in text_column_names:
            text_column_indexes.append(cell.column)

    if len(text_column_indexes) != len(text_column_names):
        workbook.close()
        raise ValueError("One or more Power BI text columns were not found in output file.")

    for row in range(2, worksheet.max_row + 1):
        for column_index in text_column_indexes:
            cell = worksheet.cell(row=row, column=column_index)
            cell.value = str(cell.value)
            cell.number_format = "@"
            cell.data_type = "s"

    workbook.save(output_file)
    workbook.close()


if __name__ == "__main__":
    main()
