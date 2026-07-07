from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ==============================================================================
# SETTINGS
# ==============================================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]

CLEAN_DATA_FOLDER = PROJECT_ROOT / "03_clean_data"

OUTPUT_FOLDER = PROJECT_ROOT / "05_matrix" / "MASTER"
OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

POWERBI_OUTPUT_FILE = OUTPUT_FOLDER / "matrix_all_amc_funds_quantity_long.xlsx"

SECURITY_MASTER_FILE = OUTPUT_FOLDER / "security_master.xlsx"

STANDARD_COLUMNS = [
    "AMC",
    "Fund_Name",
    "Portfolio_Date",
    "Month",
    "Security_Name",
    "ISIN",
    "Industry_Rating",
    "Quantity",
]

FINAL_COLUMNS = [
    "AMC",
    "Security_Name",
    "ISIN",
    "Portfolio_Date",
    "Month",
    "Month_Sort_Date",
    "Industry_Rating",
    "Fund_Name",
    "Quantity",
]

# ==============================================================================
# FORMAT HELPERS
# ==============================================================================


def format_month_text(date_series):
    """
    Converts dates to 'Jan-2025' text.
    Uses a non-breaking hyphen so Power BI
    doesn't auto-convert the value back to a date.
    """

    month_text = pd.to_datetime(date_series, errors="coerce").dt.strftime("%b-%Y")

    return month_text.str.replace("-", "\u2011", regex=False).astype("string")


def force_powerbi_text_columns(output_file):
    """
    Forces selected Excel columns to Text format
    so Power BI imports them as text.
    """

    workbook = load_workbook(output_file)
    worksheet = workbook.active

    text_columns = {
        "Portfolio_Date",
        "Month",
        "Month_Sort_Date",
    }

    column_indexes = []

    for cell in worksheet[1]:
        if cell.value in text_columns:
            column_indexes.append(cell.column)

    if len(column_indexes) != len(text_columns):
        workbook.close()
        raise ValueError("Power BI text columns not found.")

    for row in range(2, worksheet.max_row + 1):

        for column in column_indexes:

            cell = worksheet.cell(row=row, column=column)

            cell.value = str(cell.value)
            cell.number_format = "@"
            cell.data_type = "s"

    workbook.save(output_file)
    workbook.close()


# ==============================================================================
# LOAD CLEANED DATA
# ==============================================================================


def load_all_clean_data():
    """
    Reads every *_All_Funds_Cleaned.xlsx file
    from 03_clean_data and combines them.
    """

    print()
    print("=" * 90)
    print("Reading Cleaned AMC Files")
    print("=" * 90)

    cleaned_files = sorted(CLEAN_DATA_FOLDER.rglob("*_All_Funds_Cleaned.xlsx"))

    if not cleaned_files:
        raise FileNotFoundError(f"No cleaned files found in:\n{CLEAN_DATA_FOLDER}")

    frames = []
    total_rows = 0

    for file_path in cleaned_files:

        print(f"\nReading: {file_path.name}")

        df = pd.read_excel(file_path)

        missing_columns = [
            column for column in STANDARD_COLUMNS if column not in df.columns
        ]

        if missing_columns:
            raise ValueError(
                f"\n{file_path.name}\n" f"Missing columns: {missing_columns}"
            )

        df = df[STANDARD_COLUMNS].copy()

        df["Portfolio_Date"] = (
            pd.to_datetime(
                df["Portfolio_Date"],
                errors="coerce",
            )
            .dt.to_period("M")
            .dt.to_timestamp()
        )

        df = df[df["Portfolio_Date"].notna()]

        df["Quantity"] = pd.to_numeric(
            df["Quantity"],
            errors="coerce",
        ).fillna(0)

        df = df[df["Quantity"] != 0]

        frames.append(df)

        total_rows += len(df)

        print(f"Rows: {len(df)}")

    print()
    print("-" * 90)
    print(f"Files Read : {len(cleaned_files)}")
    print(f"Total Rows : {total_rows}")

    return pd.concat(
        frames,
        ignore_index=True,
    )


# ==============================================================================
# CREATE SECURITY MASTER
# ==============================================================================


def create_security_master(clean_df):
    """
    Creates a security master using ISIN as the primary key.

    If the same ISIN appears with multiple security names across AMCs,
    every name is preserved in separate columns.
    """

    print()
    print("=" * 90)
    print("Creating Security Master")
    print("=" * 90)

    security_master = (
        clean_df.groupby("ISIN")["Security_Name"]
        .apply(lambda names: sorted(set(names)))
        .reset_index()
    )

    names_df = pd.DataFrame(security_master["Security_Name"].tolist())

    names_df.columns = [f"Name_{i + 1}" for i in range(len(names_df.columns))]

    security_master = pd.concat(
        [
            security_master[["ISIN"]],
            names_df,
        ],
        axis=1,
    )

    if SECURITY_MASTER_FILE.exists():

        try:
            SECURITY_MASTER_FILE.unlink()

        except PermissionError:

            print("ERROR")
            print("Please close:")
            print(SECURITY_MASTER_FILE)

            raise

    security_master.to_excel(
        SECURITY_MASTER_FILE,
        index=False,
    )

    print()
    print("Security Master Created")
    print(f"Rows : {len(security_master)}")
    print(f"Output : {SECURITY_MASTER_FILE}")

    return security_master


# ==============================================================================
# CREATE POWER BI DATASET
# ==============================================================================


def create_powerbi_dataset(clean_df):
    """
    Creates the long-format dataset used by Power BI.
    """

    print()
    print("=" * 90)
    print("Creating Power BI Dataset")
    print("=" * 90)

    # --------------------------------------------------------------------------
    # Aggregate duplicate holdings
    # --------------------------------------------------------------------------

    powerbi_df = clean_df.groupby(
        [
            "AMC",
            "Security_Name",
            "ISIN",
            "Portfolio_Date",
            "Industry_Rating",
            "Fund_Name",
        ],
        dropna=False,
        as_index=False,
    )["Quantity"].sum()

    # --------------------------------------------------------------------------
    # Create Month columns
    # --------------------------------------------------------------------------

    powerbi_df["Month"] = format_month_text(powerbi_df["Portfolio_Date"])

    powerbi_df["Month_Sort_Date"] = powerbi_df["Portfolio_Date"].dt.strftime("%Y-%m-%d")

    powerbi_df["Portfolio_Date"] = powerbi_df["Portfolio_Date"].dt.strftime("%Y-%m-%d")

    # --------------------------------------------------------------------------
    # Reorder columns
    # --------------------------------------------------------------------------

    powerbi_df = powerbi_df[FINAL_COLUMNS]

    # --------------------------------------------------------------------------
    # Sort output
    # --------------------------------------------------------------------------

    powerbi_df = powerbi_df.sort_values(
        by=[
            "Security_Name",
            "Fund_Name",
            "Portfolio_Date",
        ],
        ascending=True,
    )

    powerbi_df.reset_index(
        drop=True,
        inplace=True,
    )

    # --------------------------------------------------------------------------
    # Save Excel
    # --------------------------------------------------------------------------

    if POWERBI_OUTPUT_FILE.exists():

        try:
            POWERBI_OUTPUT_FILE.unlink()

        except PermissionError:

            print("ERROR")
            print("Please close:")
            print(POWERBI_OUTPUT_FILE)

            raise

    powerbi_df.to_excel(
        POWERBI_OUTPUT_FILE,
        index=False,
    )

    force_powerbi_text_columns(POWERBI_OUTPUT_FILE)

    print()
    print("Power BI Dataset Created")
    print(f"Rows   : {len(powerbi_df)}")
    print(f"AMCs   : {powerbi_df['AMC'].nunique()}")
    print(f"Funds  : {powerbi_df['Fund_Name'].nunique()}")
    print(f"Stocks : {powerbi_df['ISIN'].nunique()}")
    print(f"Output : {POWERBI_OUTPUT_FILE}")

    return powerbi_df
# ==============================================================================
# MAIN
# ==============================================================================


def main():

    print("=" * 90)
    print("Building Power BI Dataset")
    print("=" * 90)

    # --------------------------------------------------------------------------
    # Load all cleaned AMC files
    # --------------------------------------------------------------------------

    clean_df = load_all_clean_data()

    # --------------------------------------------------------------------------
    # Create outputs
    # --------------------------------------------------------------------------

    create_security_master(clean_df)

    create_powerbi_dataset(clean_df)

    # --------------------------------------------------------------------------
    # Summary
    # --------------------------------------------------------------------------

    print()
    print("=" * 90)
    print("Build Completed Successfully")
    print("=" * 90)

    print(f"Total Rows : {len(clean_df)}")
    print(f"AMCs       : {clean_df['AMC'].nunique()}")
    print(f"Funds      : {clean_df['Fund_Name'].nunique()}")
    print(f"Stocks     : {clean_df['ISIN'].nunique()}")

    print()
    print("Outputs Created")
    print("------------------------------")
    print(SECURITY_MASTER_FILE)
    print(POWERBI_OUTPUT_FILE)


# ==============================================================================
# RUN
# ==============================================================================

if __name__ == "__main__":
    main()
